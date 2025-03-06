import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def parse_udemy_course(html_content, course_title):
    soup = BeautifulSoup(html_content, "html.parser")

    results = []
    current_section = None
    item_title = None  # 초기화 ✅
    index = 1  # 순서(인덱스) 추가 ✅

    for element in soup.find_all():
        if (
            element.name == "span"
            and element.get("class")
            and "section--section-title--svpHP" in element.get("class")
        ):
            current_section = element.text.strip()

        elif (
            element.name == "span"
            and element.get("class")
            and "section--item-title--EWIuI" in element.get("class")
        ):
            if current_section:
                item_title = element.text.strip()

        if (
            element.name == "span"
            and element.get("class")
            and "section--hidden-on-mobile---ITMr"
            and "section--item-content-summary--Aq9em" in element.get("class")
        ):
            if current_section:
                time_element = element.text.strip()

                # 시간이 mm:ss 형식인지 체크 (맞지 않으면 건너뛴다)
                if not re.match(r"^\d+:\d{2}$", time_element):
                    continue

                results.append(
                    {
                        "순서": index,  # ✅ 순서 추가
                        "강의 제목": course_title,
                        "섹션 타이틀": current_section,
                        "섹션 강의 아이템": (
                            item_title if item_title else "제목 없음"
                        ),  # 기본값 설정 ✅
                        "시간": time_element,
                    }
                )
                index += 1  # 순서 증가 ✅

    return pd.DataFrame(results)


def time_to_seconds(time_str):
    """
    "mm:ss" 형식의 시간을 초 단위로 변환
    """
    match = re.match(r"(\d+):(\d{2})$", time_str)
    if match:
        minutes, seconds = map(int, match.groups())
        return minutes * 60 + seconds
    return 0  # 변환 실패 시 0 반환


def highlight_rows(filename):
    """
    Excel 파일을 열어 누적 시간이 50분(3000초)을 초과하는 행에 스타일 적용.
    초과한 행 이후부터 다시 0초로 계산을 시작.
    """
    wb = load_workbook(filename)
    ws = wb.active

    orange_fill = PatternFill(
        start_color="FF8C00", end_color="FF8C00", fill_type="solid"
    )
    bold_font = Font(bold=True)

    # "시간" 컬럼이 몇 번째 열인지 찾기
    time_col = None
    for col_num, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        if col[0].value == "시간":
            time_col = col_num
            break

    if time_col is None:
        print("시간 컬럼을 찾을 수 없습니다.")
        return

    total_seconds = 0  # 누적 시간

    # 각 row별로 시간 합산 후 50분 초과하면 스타일 적용 후 시간 초기화
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        time_value = row[time_col - 1].value  # 시간 값
        time_in_seconds = time_to_seconds(time_value)

        # 시간이 변환 불가능하면 건너뛴다.
        if time_in_seconds == 0:
            continue

        total_seconds += time_in_seconds

        if total_seconds > 3000:  # 50분 초과하는 행에 스타일 적용
            for cell in row:
                cell.fill = orange_fill
                cell.font = bold_font

            total_seconds = 0  # 누적 시간 초기화 ✅ (초과한 행부터 다시 계산)

    wb.save(filename)
    print(f"✅ 50분 초과하는 행에 스타일을 적용했습니다: {filename}")


def save_to_excel(df, filename="udemy_course_analysis.xlsx"):
    df.to_excel(filename, index=False)
    highlight_rows(filename)  # 🔥 스타일 적용 함수 호출 ✅
    print(f"✅ 데이터가 {filename}에 성공적으로 저장되었습니다.")


def main():
    course_title = input("Udemy 강의 제목을 입력하세요: ")

    option = input(
        "HTML 파일 경로를 입력하려면 1, URL을 입력하려면 2, HTML 내용을 직접 붙여넣으려면 3을 입력하세요: "
    )

    if option == "1":
        file_path = input("HTML 파일 경로를 입력하세요: ")
        with open(file_path, "r", encoding="utf-8") as file:
            html_content = file.read()
    elif option == "2":
        url = input("Udemy 강의 URL을 입력하세요: ")
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        response = requests.get(url, headers=headers)
        html_content = response.text
    else:
        print("HTML 내용을 붙여넣고 마지막에 'END_HTML'을 입력하세요:")
        html_lines = []
        while True:
            line = input()
            if line == "END_HTML":
                break
            html_lines.append(line)
        html_content = "\n".join(html_lines)

    df = parse_udemy_course(html_content, course_title)

    print(f"총 {len(df)}개의 강의 아이템이 파싱되었습니다.")
    if not df.empty:
        print("파싱된 데이터 샘플:")
        print(df.head())
    else:
        print("데이터 파싱에 실패했습니다. HTML 구조를 확인해주세요.")

    if not df.empty:
        output_filename = f"{course_title.replace(' ', '-')}.xlsx"
        save_to_excel(df, output_filename)


if __name__ == "__main__":
    main()

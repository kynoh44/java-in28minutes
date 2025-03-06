import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import openpyxl
from openpyxl.styles import PatternFill, Font


def parse_udemy_course(html_content, course_title):
    soup = BeautifulSoup(html_content, "html.parser")

    results = []
    section_titles = soup.find_all("span", class_="section--section-title--svpHP")

    current_section = None
    item_title = None  # 초기화 추가 ✅
    order_number = 1  # 순서 번호 초기화 ✅

    for element in soup.find_all():
        if (
            element.name == "span"
            and element.get("class")
            and "section--section-title--svpHP" in element.get("class")
        ):
            current_section = element.text.strip()

        elif (
            element.name == "span"
            and element.get("class")
            and "section--item-title--EWIuI" in element.get("class")
        ):
            if current_section:
                item_title = element.text.strip()

        if (
            element.name == "span"
            and element.get("class")
            and "section--hidden-on-mobile---ITMr"
            and "section--item-content-summary--Aq9em" in element.get("class")
        ):
            if current_section:
                time_element = element.text.strip()
                results.append(
                    {
                        "순서": order_number,  # 순서 번호 추가 ✅
                        "강의 제목": course_title,
                        "섹션 타이틀": current_section,
                        "섹션 강의 아이템": (
                            item_title if item_title else "제목 없음"
                        ),  # 기본값 설정 ✅
                        "시간": time_element,
                    }
                )
                order_number += 1  # 순서 번호 증가 ✅

    return pd.DataFrame(results)


def convert_time_to_minutes(time_str):
    """
    mm:ss 형식의 시간 문자열을 분 단위로 변환
    mm:ss 형식이 아니면 None 반환

    Args:
        time_str (str): 시간 문자열 (mm:ss 형식)

    Returns:
        float 또는 None: 분 단위 시간 또는 형식이 맞지 않으면 None
    """
    pattern = r"(\d+):(\d+)"
    match = re.search(pattern, time_str)
    if match:
        minutes = int(match.group(1))
        seconds = int(match.group(2))
        return minutes + seconds / 60
    return None  # mm:ss 형식이 아닌 경우 None 반환


def save_to_excel(df, filename="udemy_course_analysis.xlsx"):
    """
    데이터프레임을 Excel 파일로 저장하고 특정 조건에 맞는 행에 스타일 적용

    Args:
        df (pandas.DataFrame): 저장할 데이터프레임
        filename (str): 저장할 파일 이름
    """
    # 엑셀 작성을 위한 ExcelWriter 객체 생성
    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # 데이터프레임을 Excel에 저장
    df.to_excel(writer, index=False, sheet_name="Udemy Course")

    # 워크시트 가져오기
    workbook = writer.book
    worksheet = writer.sheets["Udemy Course"]

    # 시간 누적 계산 및 강조할 행 찾기
    cumulative_time = 0
    rows_to_highlight = []
    last_valid_idx = -1

    # 시간을 순서대로 누적하여 50분을 초과하는 행 찾기
    for idx, row in df.iterrows():
        time_in_minutes = convert_time_to_minutes(row["시간"])

        # 시간 형식이 mm:ss가 아닌 경우 건너뛰기
        if time_in_minutes is None:
            continue

        # 최근 유효한 인덱스 업데이트
        last_valid_idx = idx

        # 시간 누적
        cumulative_time += time_in_minutes

        # 누적 시간이 50분을 초과하는 경우
        if cumulative_time > 50:
            rows_to_highlight.append(idx)
            print(
                f"누적 시간이 50분을 초과하는 행: {idx+2}번째 행 (누적 시간: {cumulative_time:.2f}분)"
            )
            # 시간 초기화 (다음 50분 간격을 찾기 위해)
            cumulative_time = 0

    # 마지막까지 50분을 초과하지 않은 경우 마지막 유효한 행 추가
    if cumulative_time > 0 and cumulative_time <= 50 and last_valid_idx >= 0:
        rows_to_highlight.append(last_valid_idx)
        print(
            f"마지막 세션 행: {last_valid_idx+2}번째 행 (누적 시간: {cumulative_time:.2f}분)"
        )

    # 강조할 행에 스타일 적용
    for highlight_idx in rows_to_highlight:
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(
                row=highlight_idx + 2, column=col_idx
            )  # 헤더가 1행이므로 실제 데이터는 2행부터 시작
            cell.fill = PatternFill(
                start_color="FF8C00", end_color="FF8C00", fill_type="solid"
            )
            cell.font = Font(bold=True)

    # 파일 저장
    writer.close()
    print(f"데이터가 {filename}에 성공적으로 저장되었습니다.")
    print(f"총 {len(rows_to_highlight)}개 행에 스타일이 적용되었습니다.")


def main():
    # 사용 예시
    course_title = input("Udemy 강의 제목을 입력하세요: ")

    # HTML 파일 경로 입력 또는 직접 HTML 내용 붙여넣기 옵션
    option = input(
        "HTML 파일 경로를 입력하려면 1, URL을 입력하려면 2, HTML 내용을 직접 붙여넣으려면 3을 입력하세요: "
    )

    if option == "1":
        file_path = input("HTML 파일 경로를 입력하세요: ")
        with open(file_path, "r", encoding="utf-8") as file:
            html_content = file.read()
    elif option == "2":
        url = input("Udemy 강의 URL을 입력하세요: ")
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        response = requests.get(url, headers=headers)
        html_content = response.text
    else:
        print("HTML 내용을 붙여넣고 마지막에 'END_HTML'을 입력하세요:")
        html_lines = []
        while True:
            line = input()
            if line == "END_HTML":
                break
            html_lines.append(line)
        html_content = "\n".join(html_lines)

    # 데이터 파싱
    df = parse_udemy_course(html_content, course_title)

    # 결과 확인
    print(f"총 {len(df)}개의 강의 아이템이 파싱되었습니다.")
    if not df.empty:
        print("파싱된 데이터 샘플:")
        print(df.head())
    else:
        print("데이터 파싱에 실패했습니다. HTML 구조를 확인해주세요.")

    # Excel 파일로 저장
    if not df.empty:
        output_filename = f"{course_title.replace(' ', '-')}.xlsx"
        save_to_excel(df, output_filename)


if __name__ == "__main__":
    main()
